"""
エクセルファイルの読み込み・書き出し

想定フォーマット:
  - 1行目: ヘッダー（A1: 職員番号等, B1以降: 日付 1,2,3,...）
  - 2行目以降: 職員データ（A列: 職員ID, B列以降: シフト）
"""

from io import BytesIO
from typing import Dict, List, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =========================================================
# シフト種別ごとの色設定
# =========================================================

SHIFT_COLORS = {
    "日": {"bg": "FFFFFF", "fg": "333333"},  # 白背景
    "夜": {"bg": "1e3a5f", "fg": "FFFFFF"},  # 濃紺背景・白文字
    "明": {"bg": "bbdefb", "fg": "333333"},  # 水色背景
    "公": {"bg": "c8e6c9", "fg": "333333"},  # 薄緑背景
    "希": {"bg": "fff9c4", "fg": "333333"},  # 薄黄背景
    "委": {"bg": "ffe0b2", "fg": "333333"},  # 薄橙背景
    "休": {"bg": "e0e0e0", "fg": "333333"},  # グレー背景
    "有": {"bg": "e0e0e0", "fg": "333333"},  # グレー背景
    "研": {"bg": "e1bee7", "fg": "333333"},  # 薄紫背景
}


def read_excel(file_bytes: bytes) -> Dict[str, Any]:
    """
    エクセルファイルを読み込んでJSON用の辞書を返す

    Returns:
        {
            "staff_ids": ["001", "002", ...],
            "dates": [1, 2, 3, ..., 31],
            "schedule": [["", "委", "", ...], ...],
        }
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row = 1
    start_col = None
    dates = []

    for col in range(2, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        try:
            if int(val) == 1:
                start_col = col
                break
        except (ValueError, TypeError):
            continue

    if start_col is None:
        raise ValueError("1行目で日付 '1' が入っている列を見つけられませんでした。")

    expected_day = 1
    for col in range(start_col, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        try:
            day_num = int(val)
        except (ValueError, TypeError):
            break

        if day_num != expected_day or not 1 <= day_num <= 31:
            break

        dates.append(day_num)
        expected_day += 1

    num_date_cols = len(dates)

    # 職員データを取得
    staff_ids: List[str] = []
    schedule: List[List[str]] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        staff_id = ws.cell(row_idx, 1).value
        if staff_id is None or str(staff_id).strip() == "":
            continue

        staff_ids.append(str(staff_id).strip())

        row_data = []
        for col in range(start_col, start_col + num_date_cols):
            val = ws.cell(row_idx, col).value
            if val is None:
                row_data.append("")
            else:
                row_data.append(str(val).strip())
        schedule.append(row_data)

    return {
        "staff_ids": staff_ids,
        "dates": dates,
        "schedule": schedule,
    }


def write_excel(
    staff_ids: List[str],
    dates: List,
    schedule: List[List[str]],
) -> bytes:
    """
    勤務表データからエクセルファイルを生成してバイト列で返す
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "勤務表"

    # 罫線スタイル
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="f5f5f5", end_color="f5f5f5", fill_type="solid"
    )
    header_font = Font(bold=True, size=10)
    cell_font = Font(size=10)
    center_align = Alignment(horizontal="center", vertical="center")

    # ヘッダー行
    cell = ws.cell(1, 1, "職員番号")
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align
    ws.column_dimensions["A"].width = 12

    for i, date_val in enumerate(dates):
        col = i + 2
        cell = ws.cell(1, col, date_val)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    # データ行
    for row_idx, staff_id in enumerate(staff_ids):
        # 職員ID
        cell = ws.cell(row_idx + 2, 1, staff_id)
        cell.font = cell_font
        cell.border = thin_border
        cell.alignment = center_align

        # シフトデータ
        for col_idx, shift in enumerate(schedule[row_idx]):
            cell = ws.cell(row_idx + 2, col_idx + 2, shift)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = center_align

            # 色設定
            if shift in SHIFT_COLORS:
                colors = SHIFT_COLORS[shift]
                cell.fill = PatternFill(
                    start_color=colors["bg"],
                    end_color=colors["bg"],
                    fill_type="solid",
                )
                cell.font = Font(
                    size=10,
                    color=colors["fg"],
                    bold=(shift == "夜"),
                )

    # フリーズペイン（1行目とA列を固定）
    ws.freeze_panes = "B2"

    # バイト列に変換
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
